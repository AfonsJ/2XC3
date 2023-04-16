import SPappr 
import final_project_part1
import timeit
from openpyxl import Workbook, load_workbook

# Using k values from previous experiments, runs the approx. algos with their respective "accurate" k vals
# and compares their runtime with the normal one; seeing the tradeoff of complete accuracy and runtime
def comp_runtimes():
    wrkbk = Workbook()

    runs = 10
    k_vald = 9
    k_valb = 190

    dresults  = []
    dresultsa = []
    dtime = []
    dtimea = []
    bresults = []
    bresultsa = []
    btime = []
    btimea = []
    # compare distance and time with variable relaxations; fixed 200 nodes with max val 100
    for j in range(1,10+1):
        max_val = 100
        nodes = 200
        mydict = {}
        graph = final_project_part1.create_random_complete_graph(nodes, max_val)
        for i in range(1,runs+1):
            dstarta = timeit.default_timer()
            mydict = SPappr.dijkstra_approx(graph, 0, k_vald)
            dtotala += timeit.default_timer() - dstarta
            ddista += final_project_part1.total_dist(mydict)

            dstart = timeit.default_timer()
            mydict = final_project_part1.dijkstra(graph, 0)
            dtotal += timeit.default_timer() - dstart
            ddist += final_project_part1.total_dist(mydict)
            

            bstarta = timeit.default_timer()
            mydict = SPappr.bellman_ford_approx(graph, 0, k_valb)
            btotala += timeit.default_timer() - bstarta
            bdista += final_project_part1.total_dist(mydict)

            bstart = timeit.default_timer()
            mydict = final_project_part1.bellman_ford(graph,0)
            btotal += timeit.default_timer() - bstart
            bdist += final_project_part1.total_dist(mydict)
            print("run "+str(i)+" completed. ")
        dtotala = 0
        dtotal = 0
        btotala = 0
        btotal = 0

        ddist = 0
        ddista = 0
        bdist = 0
        bdista = 0
       

        dtotala = (dtotala / runs)
        dtotal = (dtotal / runs)
        btotal = (btotal / runs)
        btotala = (btotala / runs)

        ddist = (ddist / runs)
        ddista = (ddista / runs)
        bdist = (bdist /runs)
        bdista = (bdista /runs)

        dtime.append(dtotal)
        dtimea.append(dtotala)
        btime.append(btotal)
        btimea.append(btotala)

        dresults.append(ddist)
        dresultsa.append(ddista)
        bresults.append(bdist)
        bresultsa.append(bdista)

        print("k_val "+str(j)+" completed.")

    sheetname = "Dijkstra Approx vs Main"
    worksheet = wrkbk.create_sheet(sheetname)
    worksheet.title = sheetname
    worksheet.cell(row=1, column=1, value='k_val')
    worksheet.cell(row=1, column=2, value='time_of_approx')
    worksheet.cell(row=1, column=3, value='total_dist_approx')
    worksheet.cell(row=1, column=4, value='time_of_norm')
    worksheet.cell(row=1, column=5, value='total_dist_norm')

    # i = 0
    for i in range(0,len(dtime)):
        worksheet.cell(row=i+2, column=1, value=k_vald)
        worksheet.cell(row=i+2, column=2, value=dtimea[i])
        worksheet.cell(row=i+2, column=3, value=dresultsa[i])
        worksheet.cell(row=i+2, column=4, value=dtime[i])
        worksheet.cell(row=i+2, column=5, value=dresults[i])

    sheetname = "Bellman-Ford Approx vs Main"
    worksheet = wrkbk.create_sheet(sheetname)
    worksheet.title = sheetname
    worksheet.cell(row=1, column=1, value='k_val')
    worksheet.cell(row=1, column=2, value='time_of_approx')
    worksheet.cell(row=1, column=3, value='total_dist_approx')
    worksheet.cell(row=1, column=4, value='time_of_norm')
    worksheet.cell(row=1, column=5, value='total_dist_norm')

    for i in range(0,len(dtime)):
        worksheet.cell(row=i+2, column=1, value=k_valb)
        worksheet.cell(row=i+2, column=2, value=btimea[i])
        worksheet.cell(row=i+2, column=3, value=bresultsa[i])
        worksheet.cell(row=i+2, column=4, value=btime[i])
        worksheet.cell(row=i+2, column=5, value=bresults[i])

    wrkbk.save(filename="runtime_comp.xlsx")

comp_runtimes()