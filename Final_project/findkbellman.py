import SPappr 
import final_project_part1
import timeit
from openpyxl import Workbook, load_workbook

#This function analyzes the approx. bellman-ford algo with a specific k value 
# with the normal algo and saves the data in a spreadsheet. Was used in finding the 
# k value where the approx algo's total distance matched the normal one
def get_optk_bellman():
    wrkbk = Workbook()
    runs = 10
    k_val = 0
    max_val = 80
    nodes = 100

    bresults = []
    bresultsa = []
    btime = []
    btimea = []

    for j in range(1,10+1):
        k_val = j+190 #from 10-20
        max_val = 100
        nodes = 200
        mydict = {}
        graph = final_project_part1.create_random_complete_graph(nodes, max_val)

        btotala = 0
        btotal = 0
        bdist = 0
        bdista = 0
        for i in range(1,runs+1):
            bstarta = timeit.default_timer()
            mydict = SPappr.bellman_ford_approx(graph, 0, k_val)
            btotala += timeit.default_timer() - bstarta
            bdista += final_project_part1.total_dist(mydict)

            bstart = timeit.default_timer()
            mydict = final_project_part1.bellman_ford(graph,0)
            btotal += timeit.default_timer() - bstart
            bdist += final_project_part1.total_dist(mydict)
            print("run "+str(i)+" completed. ")

        btotal = (btotal / runs)
        btotala = (btotala / runs)
        bdist = (bdist /runs)
        bdista = (bdista /runs)
        btime.append(btotal)
        btimea.append(btotala)
        bresults.append(bdist)
        bresultsa.append(bdista)
        print("k_val "+str(j)+" completed.")

    sheetname = "Bellman-Ford Approx vs Main"
    worksheet = wrkbk.create_sheet(sheetname)
    worksheet.title = sheetname
    worksheet.cell(row=1, column=1, value='k_val')
    worksheet.cell(row=1, column=2, value='time_of_approx')
    worksheet.cell(row=1, column=3, value='total_dist_approx')
    worksheet.cell(row=1, column=4, value='time_of_norm')
    worksheet.cell(row=1, column=5, value='total_dist_norm')

    for i in range(0,len(btime)):
        worksheet.cell(row=i+2, column=1, value=i+1+190)
        worksheet.cell(row=i+2, column=2, value=btimea[i])
        worksheet.cell(row=i+2, column=3, value=bresultsa[i])
        worksheet.cell(row=i+2, column=4, value=btime[i])
        worksheet.cell(row=i+2, column=5, value=bresults[i])

    wrkbk.save(filename="bellmank190.xlsx")

get_optk_bellman()