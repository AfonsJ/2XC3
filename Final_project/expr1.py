import SPappr 
import final_project_part1
import timeit
from openpyxl import Workbook, load_workbook

# Compares the runtime between Dijkstra and Bellman-Ford using a variable number of nodes (increasing graph size)
# saves the results to a spreadsheet.
def comp_reg_runtimes():
    wrkbk = Workbook()

    runs = 10
    dresults  = []
    dtime = []
    bresults = []
    btime = []

    dtotal = 0
    btotal = 0
    ddist = 0
    bdist = 0
       
    for j in range(1,10+1):
        max_val = 60
        nodes = 20*j
        mydict = {}
        graph = final_project_part1.create_random_complete_graph(nodes, max_val)

        for i in range(1,runs+1):
            dstart = timeit.default_timer()
            mydict = final_project_part1.dijkstra(graph, 0)
            dtotal += timeit.default_timer() - dstart
            ddist += final_project_part1.total_dist(mydict)
                    
            bstart = timeit.default_timer()
            mydict = final_project_part1.bellman_ford(graph,0)
            btotal += timeit.default_timer() - bstart
            bdist += final_project_part1.total_dist(mydict)
            print("run "+str(i)+" completed. ")
        

        dtotal = (dtotal / runs)
        btotal = (btotal / runs)

        ddist = (ddist / runs)
        bdist = (bdist /runs)

        dtime.append(dtotal)
        btime.append(btotal)

        dresults.append(ddist)
        bresults.append(bdist)

        print("node #"+str(j*20)+" completed.")


    sheetname = "Dijkstra Approx vs Main"
    worksheet = wrkbk.create_sheet(sheetname)
    worksheet.title = sheetname
    worksheet.cell(row=1, column=1, value='nodes')
    worksheet.cell(row=1, column=2, value='time_of_norm')
    worksheet.cell(row=1, column=3, value='total_dist_norm')

    # i = 0
    for i in range(0,len(dtime)):
        worksheet.cell(row=i+2, column=1, value=(i+1)*20)
        worksheet.cell(row=i+2, column=2, value=dtime[i])
        worksheet.cell(row=i+2, column=3, value=dresults[i])
    
    sheetname = "Bellman-Ford Approx vs Main"
    worksheet = wrkbk.create_sheet(sheetname)
    worksheet.title = sheetname
    worksheet.cell(row=1, column=1, value='nodes')
    worksheet.cell(row=1, column=2, value='time_of_norm')
    worksheet.cell(row=1, column=3, value='total_dist_norm')

    for i in range(0,len(dtime)):
        worksheet.cell(row=i+2, column=1, value=(i+1)*20)
        worksheet.cell(row=i+2, column=2, value=btime[i])
        worksheet.cell(row=i+2, column=3, value=bresults[i])

    wrkbk.save(filename="runtime_compreg.xlsx")

comp_reg_runtimes()