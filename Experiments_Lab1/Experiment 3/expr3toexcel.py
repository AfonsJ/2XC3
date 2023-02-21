from openpyxl import Workbook, load_workbook

def bubble_sort_write_to_xl(inc_index, random_list_times):
    workbook = Workbook()
    worksheet = workbook.active
    sheetname = "BubbleSort"
    # worksheet = workbook.get_active_sheet()
    worksheet.title = sheetname

    worksheet.cell(row=1, column=1, value='NUMOFSWAPS')
    worksheet.cell(row=1, column=2, value='TIME')
    #worksheet.cell(row=1, column=3, value='near_sorted_list_time')

    # i = 0
    for i in range(0,len(random_list_times)):
        worksheet.cell(row=i+2, column=1, value=inc_index)
        worksheet.cell(row=i+2, column=2, value=random_list_times[i])
        #worksheet.cell(row=i+2, column=3, value=near_sorted_list_times[i])
        inc_index+=250
    workbook.save(filename="expr3runtime.xlsx")

def insertion_sort_write_to_xl(inc_index, random_list_times):
    workbook = load_workbook('expr3runtime.xlsx')
    # worksheet = workbook.active
    sheetname = "InsertionSort"
    worksheet = workbook.create_sheet(sheetname)
    worksheet.title = sheetname

    worksheet.cell(row=1, column=1, value='NUMOFSWAPS')
    worksheet.cell(row=1, column=2, value='TIME')
    #worksheet.cell(row=1, column=3, value='near_sorted_list_time')

    # i = 0
    for i in range(0,len(random_list_times)):
        worksheet.cell(row=i+2, column=1, value=inc_index)
        worksheet.cell(row=i+2, column=2, value=random_list_times[i])
        #worksheet.cell(row=i+2, column=3, value=near_sorted_list_times[i])
        inc_index+=250
    workbook.save("expr3runtime.xlsx")

def selection_sort_write_to_xl(inc_index, random_list_times):
    workbook = load_workbook('expr3runtime.xlsx')
    # worksheet = workbook.active
    sheetname = "SelectionSort"
    worksheet = workbook.create_sheet(sheetname)
    worksheet.title = sheetname

    worksheet.cell(row=1, column=1, value='NUMOFSWAPS')
    worksheet.cell(row=1, column=2, value='TIME')
    #worksheet.cell(row=1, column=3, value='near_sorted_list_time')

    # i = 0
    for i in range(0,len(random_list_times)):
        worksheet.cell(row=i+2, column=1, value=inc_index)
        worksheet.cell(row=i+2, column=2, value=random_list_times[i])
        #worksheet.cell(row=i+2, column=3, value=near_sorted_list_times[i])
        inc_index+=250
    workbook.save("expr3runtime.xlsx")