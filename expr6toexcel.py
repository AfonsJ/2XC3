from openpyxl import Workbook, load_workbook


def quick_sort_write_to_xl(inc_index, times):
    workbook = Workbook()
    worksheet = workbook.active
    sheetname = "QuickSort"
    # worksheet = workbook.get_active_sheet()
    worksheet.title = sheetname

    worksheet.cell(row=1, column=1, value='LIST LENGTH')
    worksheet.cell(row=1, column=2, value='TIME')
    list_length = inc_index
    #worksheet.cell(row=1, column=3, value='near_sorted_list_time')

    # i = 0
    for i in range(0,len(times)):
        worksheet.cell(row=i+2, column=1, value=list_length)
        worksheet.cell(row=i+2, column=2, value=times[i])
        #worksheet.cell(row=i+2, column=3, value=near_sorted_list_times[i])
        list_length+=inc_index
    workbook.save(filename="expr6runtime.xlsx")

def dual_quick_sort_write_to_xl(inc_index, times):
    workbook = load_workbook('expr6runtime.xlsx')
    # worksheet = workbook.active
    sheetname = "Dual QuickSort"
    worksheet = workbook.create_sheet(sheetname)
    worksheet.title = sheetname

    worksheet.cell(row=1, column=1, value='LIST LENGTH')
    worksheet.cell(row=1, column=2, value='TIME')
    list_length = inc_index
    #worksheet.cell(row=1, column=3, value='near_sorted_list_time')

    # i = 0
    for i in range(0,len(times)):
        worksheet.cell(row=i+2, column=1, value=list_length)
        worksheet.cell(row=i+2, column=2, value=times[i])
        #worksheet.cell(row=i+2, column=3, value=near_sorted_list_times[i])
        list_length+=inc_index
    workbook.save("expr6runtime.xlsx")
